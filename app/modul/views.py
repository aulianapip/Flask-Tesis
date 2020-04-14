from flask import render_template, request #buat memanggil .html
from app import app
from openpyxl import load_workbook
import pandas as pd
import os
import numpy as np
import dbmodel as database
import xlrd
import mysql.connector
from mysql.connector import errorcode
import json
import plotly
from nltk.tokenize import word_tokenize
from Sastrawi.Stemmer.StemmerFactory import StemmerFactory
from Sastrawi.StopWordRemover.StopWordRemoverFactory import StopWordRemoverFactory
from vectorizer import cluster_paragraphs
from random import shuffle
remover= StopWordRemoverFactory().create_stop_word_remover()
stemmer= StemmerFactory().create_stemmer()
factory = StopWordRemoverFactory()
stopword = factory.create_stop_word_remover()




@app.route('/upload', methods= ['GET','POST'])
def upload():
    return render_template("upload_data.html")

@app.route('/upload_hasil', methods= ['GET','POST'])
def upload_hasil():
    if request.method == 'POST':
        a=request.files['file'] #variabel untuk nyimpan file
        b=request.form['sheet']

        a.save(os.path.join('app/upload_data', 'DATA EXCEL.xlsx')) #wadah untuk setiap kita nge-load, hasilnya disimpan disitu
        wb = load_workbook(filename='app/upload_data/DATA EXCEL.xlsx')
        sheet_ranges = wb[request.form['sheet']]
        data = pd.DataFrame(sheet_ranges.values)

    return render_template('hasil_upload.html', tables=[data.to_html(classes='Alumni')],sheet_ranges=b)




@app.route('/hasil_select', methods=['GET','POST'])
def hasil_select():
    if request.method == 'POST':
        select1 = request.form["select1"]
        select2 = request.form["select2"]
        selectkolom = request.form["selectkolom"]
        namakolom = request.form["namakolom"]

        wb = load_workbook(filename='app/upload_data/DATA EXCEL.xlsx')
        sheet_ranges = wb[request.form['sheet']]
        data = pd.DataFrame(sheet_ranges.values)

        row1 = int(select1)
        row2 = int(select2)

        cols = selectkolom.split(",") #memisahkan kolom dengan koma
        cols = list(map(int,cols)) #convert to list
        xname = namakolom.split(",") #memisahkan nama kolom dengan koma
        data = data[row1:row2][cols]
        data.columns = ['Judul','Date']

        header = {}
        for index, head in enumerate(xname):
            header[str(index)] = head
        data['Date'] = pd.DatetimeIndex(data['Date']).year  
        grouped = data.groupby('Date')
        # a=grouped.get_group(2012)
        b=grouped.get_group(2013)
        c=grouped.get_group(2014)
        d=grouped.get_group(2015)
        # e=grouped.get_group(2016)
        # f=grouped.get_group(2017)`
        data=b
        data1=c

        dbmodel = database.DBModel() #memanggil file model dimodel class DBModel
        result_insert_table = dbmodel.insert_cleaning_data("DataTA","data 2013",data)
        result_insert_table = dbmodel.insert_cleaning_data("DataTA","data 2014",data1)
        result_insert_header = dbmodel.insert_header("DataTA","judulnya",header)

    return render_template('tampil_hasil_select.html', tables1=[data.to_html(classes='Alumni')],tables2=[data1.to_html(classes='Alumni')])

@app.route('/dataframe', methods=['GET','POST'])
def Dataframe():
    if request.method == 'POST':
        dbmodel = database.DBModel()
        token = dbmodel.get_data_all("DataTA","datanya")

        data_s=[]
        for i in token :
            isi = i.values()
            isi_judul = isi[1]
            data_baru2 = isi_judul.lower()
            data_s.append(stopword.remove(data_baru2))

        data = data_s
        shuffle(data)

        cluster_paragraphs(data, num_clusters=2)
        clusters = cluster_paragraphs(data, num_clusters=2)
        data = pd.DataFrame(clusters)

        

    return render_template('dataframe.html', tables=[data.to_html(classes='Alumni')])

@app.route('/coba', methods=['GET','POST'])
def Coba():
    if request.method == 'POST':
        dbmodel = database.DBModel()
        token = dbmodel.get_data_spe("DataTA","data 2013")
        token1= dbmodel.get_data_spe("DataTA","data 2014")

        data_s=[]
        for i in token :
            isi = i.values()
            isi_judul = isi[0]
            data_baru2 = isi_judul.lower()
            data_s.append(stopword.remove(data_baru2))

        data = data_s
        print data
        

        cluster_paragraphs(data, num_clusters=2)
        clusters = cluster_paragraphs(data, num_clusters=2)
        # da = np.concatenate(clusters)
        data=pd.DataFrame(clusters, index=['c1','c2'])
        data=data.T
        dbmodel = database.DBModel()
        result_insert_table = dbmodel.insert_cleaning_data("DataTA","data cluster 2013",data)
        # data1 = pd.DataFrame(clusters[0])
        # data2 = pd.DataFrame(clusters[1])
        # c = 'Group 1:========\n\n-----\n'.join(t for t in clusters[0])
        # d = 'Group 2:========\n\n-----\n'.join(t for t in clusters[1])
    return render_template('hasil_cluster.html', tables=[data.to_html(classes='Alumni')])

@app.route('/upload_mysql', methods= ['GET','POST'])
def upload_mysql():
    if request.method == 'POST':
        a=request.files['file'] #variabel untuk nyimpan file
        b=request.form['sheet']

        a.save(os.path.join('app/upload_data', 'DATA EXCEL.xlsx')) #wadah untuk setiap kita nge-load, hasilnya disimpan disitu
        wb = load_workbook(filename='app/upload_data/DATA EXCEL.xlsx')
        sheet_ranges = wb[request.form['sheet']]
        data = pd.DataFrame(sheet_ranges.values)

        list= xlrd.open_workbook("app/upload_data/DATA EXCEL.xlsx")
        sheet= list.sheet_by_index(0)

        database = mysql.connector.connect (host="localhost" , user="root" , passwd="" ,db="Flask")
        cursor = database.cursor()
        cursor.execute("USE Flask")
        query= """INSERT INTO data(id, nim, nama, judul, tanggal, angkatan) VALUES (%s, %s, %s, %s, %s, %s)"""



        for r in range(1,sheet.nrows):
            id = sheet.cell(r,0).value
            nim = sheet.cell(r,1).value
            nama=sheet.cell(r,2).value
            judul=sheet.cell(r,3).value #>>>>> HERE THE PROBLEM the Imported Value <<<<
            tanggal=sheet.cell(r,4).value
            angkatan=sheet.cell(r,5).value
            values = (id, nim, nama, judul, tanggal, angkatan)

            cursor.execute(query,values)



        cursor.close();
        database.commit()

        database.close()

    return render_template('load_data.html', tables=[data.to_html(classes='Alumni')],sheet_ranges=b)

@app.route('/load_data', methods=['GET','POST'])
def Load_data():
    
    return render_template('load_data.html')

@app.route('/load_data5', methods=['GET','POST'])
def Load_data2():
    if request.method == 'POST':
        a=request.files['file'] #variabel untuk nyimpan file
        b=request.form['sheet']

        a.save(os.path.join('app/upload_data', 'DATA EXCEL.xlsx')) #wadah untuk setiap kita nge-load, hasilnya disimpan disitu
        wb = load_workbook(filename='app/upload_data/DATA EXCEL.xlsx')
        sheet_ranges = wb[request.form['sheet']]
        data = pd.DataFrame(sheet_ranges.values)

    return render_template('load_data2.html', tables=[data.to_html(classes='Alumni')],sheet_ranges=b)

@app.route('/load_data3', methods=['GET','POST'])
def Load_data3():
    if request.method == 'POST':
        select1 = request.form["select1"]
        select2 = request.form["select2"]
        selectkolom = request.form["selectkolom"]
        namakolom = request.form["namakolom"]

        wb = load_workbook(filename='app/upload_data/DATA EXCEL.xlsx')
        sheet_ranges = wb[request.form['sheet']]
        data = pd.DataFrame(sheet_ranges.values)

        row1 = int(select1)
        row2 = int(select2)

        cols = selectkolom.split(",") #memisahkan kolom dengan koma
        cols = list(map(int,cols)) #convert to list
        xname = namakolom.split(",") #memisahkan nama kolom dengan koma
        data = data[row1:row2][cols]
        data.columns = [xname]

        header = {}
        for index, head in enumerate(xname):
            header[str(index)] = head




        dbmodel = database.DBModel() #memanggil file model dimodel class DBModel
        result_insert_table = dbmodel.insert_cleaning_data("DataTA","datanya",data)
        result_insert_header = dbmodel.insert_header("DataTA","judulnya",header)

    return render_template('load_data3.html', tables=[data.to_html(classes='Alumni')])

@app.route('/index', methods=['GET','POST'])
def Index():
    
    return render_template('index.html')    

@app.route('/text_mining', methods=['GET','POST'])
def Text_mining():
    if request.method == 'POST':
        dbmodel = database.DBModel()
        token = dbmodel.get_data_all("DataTA","datanya")

        data_s=[]
        for i in token :
            isi = i.values()
            isi_judul = isi[1]
            data_baru2 = isi_judul.lower()
            word_token2 = word_tokenize(data_baru2)
            data_s.append((word_token2))

    
        data = pd.DataFrame(data_s)
        head = []
        for j in data.columns:
            head_string = "T" + str(j)
            head.append(head_string)

        data.columns = head
        dbmodel = database.DBModel()  # memanggil file model dimodel class DBModel
        result_insert_table = dbmodel.insert_tokenisasi_data("DataTA", "Tokenisasi", data)


    return render_template('text_mining.html', tables=[data.to_html(classes='Alumni')])  

@app.route('/text_filtering', methods=['GET','POST'])
def Text_filtering():
    if request.method == 'POST':
        dbmodel = database.DBModel()
        filter = dbmodel.get_data_all("DataTA","Tokenisasi")

        data_x = []
        for x in filter:
            fils = x.values()
            isi_fils = fils[:-1]
            
            data_filter = []
            for z in isi_fils:
                if z <> None:
                    b = (z.encode("ascii","ignore"))
                    stop_w = remover.remove(b)
                    if stop_w <> "":
                        data_filter.append(stop_w)
            data_x.append(data_filter)

            data = pd.DataFrame(data_x)
            head = []
            for j in data.columns:
                head_string = "T" + str(j)
                head.append(head_string)

            data.columns = head
            dbmodel = database.DBModel()  # memanggil file model dimodel class DBModel
            result_insert_table = dbmodel.insert_filtering_data("DataTA", "Filtering", data)

    return render_template('text_filtering.html', tables=[data.to_html(classes='Alumni')]) 
@app.route('/text_stemming', methods=['GET','POST'])
def Text_stemming():
    if request.method == 'POST':
        dbmodel = database.DBModel()
        stem = dbmodel.get_data_all("DataTA","Filtering")

        data_stem = []
        for x in stem:
            stemm = x.values()
            isi_stem = stemm[:-1]
            data_stemming = []
            for z in isi_stem:
                if z <> None:
                    b = (z.encode("ascii", "ignore"))
                    stem_w = stemmer.stem(b)
                    if stem_w <> "":
                        data_stemming.append(stem_w)
            data_stem.append(data_stemming)

            data = pd.DataFrame(data_stem)
            head = []
            for j in data.columns:
                head_string = "T" + str(j)
                head.append(head_string)

            data.columns = head
            dbmodel = database.DBModel()  # memanggil file model dimodel class DBModel
            result_insert_table = dbmodel.insert_stemming_data("DataTA", "Stemming", data)

        return render_template('text_stemming.html', tables=[data.to_html(classes='Alumni')])

@app.route('/tokenizing', methods=['GET','POST'])
def Tokenizing():
    if request.method == 'POST':
        dbmodel = database.DBModel()
        token = dbmodel.get_data_all("DataTA","datanya")

        data_s=[]
        for i in token :
            isi = i.values()
            isi_judul = isi[1]
            data_baru2 = isi_judul.lower()
            word_token2 = word_tokenize(data_baru2)
            data_s.append((word_token2))

    
        data = pd.DataFrame(data_s)
        head = []
        for j in data.columns:
            head_string = "T" + str(j)
            head.append(head_string)

        data.columns = head
        dbmodel = database.DBModel()  # memanggil file model dimodel class DBModel
        result_insert_table = dbmodel.insert_tokenisasi_data("DataTA", "Tokenisasi", data)


    return render_template('tokenizing.html', tables=[data.to_html(classes='Alumni')])


@app.route('/filtering', methods=['GET','POST'])
def Filtering():
    if request.method == 'POST':
        dbmodel = database.DBModel()
        filter = dbmodel.get_data_all("DataTA","Tokenisasi")

        data_x = []
        for x in filter:
            fils = x.values()
            isi_fils = fils[:-1]
            
            data_filter = []
            for z in isi_fils:
                if z <> None:
                    b = (z.encode("ascii","ignore"))
                    stop_w = remover.remove(b)
                    if stop_w <> "":
                        data_filter.append(stop_w)
            data_x.append(data_filter)

            data = pd.DataFrame(data_x)
            head = []
            for j in data.columns:
                head_string = "T" + str(j)
                head.append(head_string)

            data.columns = head
            dbmodel = database.DBModel()  # memanggil file model dimodel class DBModel
            result_insert_table = dbmodel.insert_filtering_data("DataTA", "Filtering", data)

    return render_template('filtering.html', tables=[data.to_html(classes='Alumni')])

@app.route('/stemming', methods=['GET','POST'])
def Stemming():
    if request.method == 'POST':
        dbmodel = database.DBModel()
        stem = dbmodel.get_data_all("DataTA","Filtering")

        data_stem = []
        for x in stem:
            stemm = x.values()
            isi_stem = stemm[:-1]
            data_stemming = []
            for z in isi_stem:
                if z <> None:
                    b = (z.encode("ascii", "ignore"))
                    stem_w = stemmer.stem(b)
                    if stem_w <> "":
                        data_stemming.append(stem_w)
            data_stem.append(data_stemming)

            data = pd.DataFrame(data_stem)
            head = []
            for j in data.columns:
                head_string = "T" + str(j)
                head.append(head_string)

            data.columns = head
            dbmodel = database.DBModel()  # memanggil file model dimodel class DBModel
            result_insert_table = dbmodel.insert_stemming_data("DataTA", "Stemming", data)

        return render_template('stemming.html', tables=[data.to_html(classes='Alumni')])

@app.route('/grafik', methods=['GET','POST'])
def Grafik():
    if request.method == 'POST':
        dbmodel = database.DBModel()
        token = dbmodel.get_data_cluster("DataTA","data cluster 2013")
        token1 = dbmodel.get_data_cluster1("DataTA","data cluster 2013")
        data_s=[]
        for i in token :
            isi = i.values()
            isi_judul = isi[0]
            data_baru2 = isi_judul.lower()
            data_s.append(stopword.remove(data_baru2))

        data = data_s
        a = len(data)
        print a

        data_s1=[]
        for j in token1 :
            isi1 = j.values()
            isi_judul1 = isi1[0]
            data_baru3 = isi_judul1.lower()
            data_s1.append(stopword.remove(data_baru3))

        data2 = data_s1
        b = len(data2)
        print b

    rng = pd.date_range('1/1/2011', periods=7500, freq='H')
    ts = pd.Series(np.random.randn(len(rng)), index=rng)

    graphs = [
        dict(
            data=[
                dict(
                    x=[2013, 2014, 2015],
                    y=[a, 20, 30],
                    type='scatter',
                    mode='lines',
                    name = 'c1'
                ),
                 dict(
                    x=[2013, 2014, 2015],
                    y=[b, 25, 35],
                    type='scatter',
                    mode='lines',
                    name = 'Sistem Produksi'
                )

            ],
            layout=dict(
                title='Trend Data Skripsi Teknik Industri'
            )
        ),

        

        dict(
            data=[
                dict(
                    x=ts.index,  # Can use the pandas data structures directly
                    y=ts
                )
            ]
        )
    ]

    # Add "ids" to each of the graphs to pass up to the client
    # for templating
    ids = [format(i) for i, _ in enumerate(graphs)]

    # Convert the figures to JSON
    # PlotlyJSONEncoder appropriately converts pandas, datetime, etc
    # objects to their JSON equivalents
    graphJSON = json.dumps(graphs, cls=plotly.utils.PlotlyJSONEncoder)

    return render_template('grafik.html',
                           ids=ids,
                           graphJSON=graphJSON)
    






